from __future__ import annotations

import sys, typing

import openpyxl as pxl
from openpyxl import styles as pxl_sty 
from openpyxl.worksheet import worksheet as pxl_sht 
from openpyxl.worksheet import cell_range as pxl_rng

sys.path.insert(0, '..\..')

from db_controllers import _controller_base
from db_controllers.entity_capsules import _capsule_base


from db_controllers.xl.sheet import io_sht as io_sht
from db_controllers.xl.validation import validation_sht as io_val



CNTR = typing.TypeVar("CNTR", bound=_controller_base.ControllerBase)
CT = typing.TypeVar("CT", bound=_capsule_base.CapsuleBase)

class IoWorkbook():
  def __init__(self, 
               subControllerKey: _controller_base.ControllerKeyEnum,
               capsuleTypes: typing.List[type[CT]],
               validationLocators: typing.List[(str, str)]) -> None:
    self.subControllerKey = subControllerKey
    self._capsuleTypes = capsuleTypes
    self.validations = io_val.ValidationSheet(
               subControllerKey = subControllerKey,
               validationLocators = validationLocators)
    self.subControllerDict: typing.Dict = {}
    self.ioShts: typing.OrderedDict[str, io_sht.IoWorkSheet] = {}
    self._initIoShts()

    self.path = None
    self.wkb: pxl.Workbook = None

  def _initIoShts(self):
    for capsuleType in self._capsuleTypes:
      ioWorkSheet = io_sht.IoWorkSheet(
                subControllerKey = self.subControllerKey,
                capsuleType = capsuleType,
                validations=self.validations)
      self.ioShts[ioWorkSheet.name] = ioWorkSheet

  def save(self,
           path: str = ""):
    if len(self.path) == 0 and len(path) == 0:
      raise Exception("Can't save workbook without path")
    elif len(self.path) == 0 and len(path) > 0:
      self.path = path
    self.wkb.save(self.path)

  def setupDownload(self,
                    controllerDict: typing.Dict,
                    path: str) -> None:
    self.path = path
    self.controllerDict = controllerDict
    self.wkb = pxl.Workbook()
    sheetNamesToDelete = self.wkb.sheetnames
    self.validations.setup(wkb = self.wkb,
                           controllerDict = controllerDict)
    for ioSht in self.ioShts.values():
      ioSht.setup(wkb = self.wkb,
                  controllerDict = self.controllerDict)
    # remove superfluous worksheets
    for sheetNameToDelete in sheetNamesToDelete:
       sheetToDelete = self.wkb[sheetNameToDelete]
       self.wkb.remove(sheetToDelete)
    # move the validation sheet to the end and hide
    move = len(self.wkb.worksheets) - \
           self.wkb.worksheets.index(self.validations.sht) \
           - 1
    self.wkb.move_sheet(self.validations.sht, move)
    self.validations.sht.sheet_state = 'hidden'
    self.wkb.active = 0
    self.save(path = self.path)

  def loadAndIdentifyUploadWkb(self,
                               path: str):
    try:
      self.wkb = pxl.load_workbook(filename=path)
    except:
      raise Exception(f"Can't load ioWkb: {path}\n" +\
                      f"No such file or directory.")
    # FIXME: identify validation sheet
    for ioSht in self.ioShts.values():
      ioSht.identify(wkb = self.wkb)

  def toDict(self) -> typing.Dict[str, any]:    
    capsulesDict: typing.Dict[str, any] = dict[str, any]()
    for ioSht in self.ioShts.values():
      capsuleDict = ioSht.toDict()
      capsulesDict = capsulesDict | capsuleDict
    result: typing.Dict[str, any] = dict[str, any]()
    subControllerKey = self.subControllerKey.value
    result[subControllerKey] = capsulesDict
    return result