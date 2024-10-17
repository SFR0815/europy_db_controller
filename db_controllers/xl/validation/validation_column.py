from __future__ import annotations

import sys, typing

import sqlalchemy

import openpyxl as pxl
from openpyxl import styles as pxl_sty 
from openpyxl.workbook import defined_name as pxl_rng_nm
from openpyxl.formatting import rule as pxl_frm_rule
from openpyxl.worksheet import cell_range as pxl_rng
from openpyxl.worksheet import worksheet as pxl_sht
from openpyxl.worksheet import datavalidation as pxl_dv
from openpyxl.cell import cell as pxl_cell

sys.path.insert(0, '..\..\..')

from db_controllers import _controller_base
from db_controllers.entity_capsules import _capsule_base
from db_controllers.xl.sheet import utils as sht_ut

C = typing.TypeVar("C", bound=_capsule_base.CapsuleBase)

class ValidationColumn():
  _labelRow = 3
  _startRow = 4
  _patternFill = pxl_sty.PatternFill(start_color="F6F2A0", end_color="F6F2A0", fill_type = 'solid')
  _borderStyle1 = pxl_sty.Side(style='thin', color='000000')
  _borderStyle2 = pxl_sty.Side(style='dotted', color='000000')
  _borderLabel = pxl_sty.Border(left=_borderStyle1, right=_borderStyle1,
                           top=_borderStyle1, bottom=_borderStyle1,
                           vertical=None, horizontal=None)
  _borderData = pxl_sty.Border(left=_borderStyle1, right=_borderStyle1,
                           top=None, bottom=_borderStyle2,
                           vertical=None, horizontal=None)
  _labelAlignment = pxl_sty.Alignment(textRotation=90, horizontal = 'center')

  @classmethod
  def getColumnLabel(self,
                     validationLocator: typing.Tuple[str, str]
                     ) -> str:
    validationSubControllerKey, validationCapsuleKey = validationLocator
    return f"{validationSubControllerKey}.{validationCapsuleKey}"

  def __init__(self,
               columnNumber: int, 
               subControllerKey: _controller_base.ControllerKeyEnum,
               validationLocator: typing.Tuple[str, str]
               ) -> None:
    self.subControllerKey = subControllerKey
    self.validationLocator = validationLocator
    self._colNo: int = columnNumber
    self.sht: pxl_sht.Worksheet = None
    self.wkb: pxl.Workbook = None
    self.validationDict: typing.Dict = None
  
  @property
  def isInternalValidation(self) -> bool:
    return self._colNo == -1
  @property
  def columnNumber(self) -> int:
    return self._colNo
  @property
  def label(self) -> str:
    return self.__class__.getColumnLabel(
                      validationLocator = self.validationLocator)
  @property
  def validationSubControllerId(self) -> str:
    return self.validationLocator[0]
  @property
  def validationCapsuleId(self) -> str:
    return self.validationLocator[1]
  @property
  def columnLabelRow(self) -> int:
    return self._labelRow
  @property
  def labelCell(self) -> pxl_cell.Cell:
    return self.sht.cell(self.columnLabelRow, self.columnNumber)
  @property
  def validationRangeDelimiters(self) -> typing.Dict[str, int]:
    result = dict[str, int]()
    result['min_row'] = self.columnLabelRow + 1
    result['max_row'] = self.columnLabelRow + (1 if len(self.validationDict) == 0 else len(self.validationDict))
    result['min_col'] = self.columnNumber
    result['max_col'] = self.columnNumber
    return result
  @property
  def validationCellRange(self) -> pxl_rng.CellRange:
    return pxl_rng.CellRange(**self.validationRangeDelimiters)
  @property
  def quotedSheetTitle(self):
    return pxl.utils.cell.quote_sheetname(self.sht.title)
  @property
  def localRangeName(self) -> str:
    label = self.__class__.getColumnLabel(validationLocator = self.validationLocator)
    return label.replace('.', '_')
  @property
  def globalRangeName(self) -> str:
    return f"{self.quotedSheetTitle}!{self.localRangeName}"
  @property 
  def fullRangeAddress(self) -> str:
    coord = pxl.utils.absolute_coordinate(self.validationCellRange.coord)
    return f"{self.quotedSheetTitle}!{coord}"

  def dataValidation(self,
                     quotedSheetTitle: str = None,
                     cellRange: pxl_rng.CellRange = None) -> pxl_dv.DataValidation:
    def getValidationFormula() -> str:      
      if self.isInternalValidation:
        if quotedSheetTitle is None or cellRange is None:
          raise Exception("[ValidationColumn - dataValidation] Must provide " +\
                          "'quotedSheetTitle' and 'validationRangeName'" +\
                          "if the validation is within the subcontroller of the data workbook.")
        coord = pxl.utils.absolute_coordinate(cellRange.coord)
        return f"{quotedSheetTitle}!{coord}"
      else: 
        return f'{self.fullRangeAddress}'
    return pxl_dv.DataValidation(type='list', 
                                formula1=getValidationFormula(), 
                                allow_blank=True,
                                error = 'Please select a value from the list.',
                                errorTitle = 'Non-permissible data entry.',
                                showErrorMessage=True)  
  # FIXME provide def for CellIsRule if validation condition is not met

  def setup(self,
            validationDict: typing.Dict,
            wkb: pxl.Workbook,
            sht: pxl_sht.Worksheet) -> None:
    self.wkb = wkb
    self.sht = sht
    self.validationDict = validationDict
    self.labelCell.value = self.label
    self.labelCell.border = self._borderLabel
    self.labelCell.fill = self._patternFill
    self.labelCell.alignment = self._labelAlignment
    dictCount: int = 0
    if len(self.validationDict) == 0: return
    for validationCellTuple in self.sht.iter_rows(**self.validationRangeDelimiters):
      validationValue = self.validationDict[dictCount]['name']
      validationCell: pxl_sht.Cell = validationCellTuple[0]
      validationCell.value = validationValue
      validationCell.border = self._borderData
      validationCell.fill = self._patternFill
      dictCount += 1
    definedName = pxl_rng_nm.DefinedName(name = self.localRangeName,
                                         attr_text = self.fullRangeAddress)
    self.sht.defined_names.add(definedName)




