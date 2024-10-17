from __future__ import annotations

import sys, typing

from openpyxl import styles as pxl_sty 
from openpyxl.workbook import defined_name as pxl_rng_nm
from openpyxl.worksheet import worksheet as pxl_sht
from openpyxl.worksheet import datavalidation as pxl_dv
from openpyxl.worksheet import cell_range as pxl_rng
from openpyxl.cell import cell as pxl_cell
from openpyxl.formatting import rule as pxl_frm_rule

sys.path.insert(0, '..\..\..')

from db_controllers import _controller_base

from db_controllers.xl.sheet import data_column

if typing.TYPE_CHECKING:
  from db_controllers.xl.sheet import row_control
  from db_controllers.xl.sheet import col_control

DELETE_CONTROL_LABEL = 'delete_control'
class DeleteControlColumn(data_column.DataColumn):
  _patternFill = pxl_sty.PatternFill(start_color="FF0000", end_color="FF0000", fill_type = 'solid')
  _borderStyle1 = pxl_sty.Side(style='thick', color='000000')
  _borderStyle2 = pxl_sty.Side(style='dotted', color='000000')
  _border = pxl_sty.Border(left=_borderStyle1, right=_borderStyle1,
                           top=None, bottom=_borderStyle2,
                           vertical=None, horizontal=_borderStyle2)
  _labelAlignment = pxl_sty.Alignment(textRotation=90, horizontal = 'center')


  _deletePatternFill = pxl_sty.PatternFill(start_color="FF0000", end_color="FF0000", fill_type = 'solid')
  _deleteFont =pxl_sty.Font(bold = True)
  _deleteCellFlag = '"X"'
  _deleteFormatRule = pxl_frm_rule.CellIsRule(operator='equal', formula=[_deleteCellFlag],
                                                stopIfTrue=True, fill=_deletePatternFill,
                                                font=_deleteFont)
  
  _deleteValidation = pxl_dv.DataValidation(type='custom', operator='equal', formula1='"X"')
  _deleteValidation.error = 'Value must be empty or exaclty equal to an uppercase "X".'
  _deleteValidation.errorTitle = 'Bad delete flag'

  def __init__(self,
               subControllerKey: _controller_base.ControllerKeyEnum,
               columnNumber: int,
               rowControl: 'row_control.RowControl',
               colControl: 'col_control.ColControl'
               ) -> None:
    super().__init__(label = DELETE_CONTROL_LABEL,
                     subControllerKey = subControllerKey,
                     validation = None,
                     columnNumber = columnNumber,
                     rowControl = rowControl,
                     colControl = colControl, 
                     unique = False,
                     sqlalchemyDataType = None)

  def setFormatsAndValidations(self,
                               sht: pxl_sht.Worksheet) -> None:
    self.sht = sht
    definedName = pxl_rng_nm.DefinedName(name = self.localDataRangeName,
                                         attr_text = self.fullDataRangeAddress)
    self.sht.defined_names.add(definedName)
    self.labelCell.border = self._colControl._border
    self.labelCell.fill = self._patternFill
    self.labelCell.alignment = self._labelAlignment
    self.labelCell.value = self.label
    if self.subControllerKey == _controller_base.ControllerKeyEnum.ASSET_PRICING:
      usedRows = self._rowControl.lastDataRow + self._numberOfRowsPricing
    else:
      usedRows = self._rowControl.lastDataRow + self._numberOfRows

    for rowCellTuple in self.sht.iter_rows(min_row=self.columnLabelRow + 1, 
                                      max_row=usedRows, 
                                      min_col=self.columnNumber,
                                      max_col=self.columnNumber):
      rowCell: pxl_sht.Cell = rowCellTuple[0]
      rowCell.fill = super()._patternFill
      rowCell.border = self._border
    dataValidation = pxl_dv.DataValidation(type='list', 
                                           formula1='"X"', 
                                           allow_blank=True,
                                           error = 'Value must be empty or exactly equal to an uppercase "X" - or empty.',
                                           errorTitle = 'Bad delete flag',
                                           showErrorMessage=True)
    ws = self.sht
    dataValidation.add(self.dataRange.coord)
    self.sht.add_data_validation(dataValidation)
    self.sht.conditional_formatting.add(self.dataRange.coord, self._deleteFormatRule)



    

