from __future__ import annotations

import typing, enum

import openpyxl as pxl
from openpyxl import utils as pxl_utl
from openpyxl import styles as pxl_sty 
from openpyxl.worksheet import worksheet as pxl_sht 
from openpyxl.cell import cell as pxl_cell
from openpyxl.worksheet import cell_range as pxl_rng

def hasSheetOfName(wkb: pxl.Workbook,
                   shtName: str) -> bool:
  try: 
    _ = wkb[shtName]
    return True
  except:
    return False

class CellAddressType(enum.Enum):
  FIXED_COLUMN = 0
  FIXED_ROW = 1
  FIXED = 2
  RELATIVE = 3  

def getCellAddress(row: int, col: int, 
                   addressType: CellAddressType = CellAddressType.RELATIVE) -> str:
  columnId = pxl_utl.get_column_letter(col)
  if addressType == CellAddressType.FIXED or addressType == CellAddressType.FIXED_COLUMN:
    columnId = '$' + columnId
  rowId = str(row)
  if addressType == CellAddressType.FIXED or addressType == CellAddressType.FIXED_ROW:
    rowId = '$' + rowId
  return columnId + rowId
def getRangeAddress(luRow: int, luCol: int,
                    rlRow: int, rlCol: int, 
                    addressType: CellAddressType = CellAddressType.RELATIVE) -> str:
  luColLetter = pxl_utl.get_column_letter(luCol)
  rlColLetter = pxl_utl.get_column_letter(rlCol)
  if addressType == CellAddressType.FIXED or addressType == CellAddressType.FIXED_COLUMN:
    luColLetter = '$' + luColLetter
    rlColLetter = '$' + rlColLetter
  luRowId = str(luRow)
  rlRowId = str(rlRow)
  if addressType == CellAddressType.FIXED or addressType == CellAddressType.FIXED_ROW:
    luRowId = '$' + luRowId
    rlRowId = '$' + rlRowId
  return f"{luColLetter}{luRowId}:{rlColLetter}{rlRowId}"

def updateBorderParameter(border: pxl_sty.Border,
                     paramValue: any,
                     paramName: str) -> pxl_sty.Border:
  paramDict: typing.Dict = {}
  paramDict['left'] = border.left
  paramDict['right'] = border.right
  paramDict['top'] = border.top
  paramDict['bottom'] = border.bottom
  paramDict['diagonal'] = border.diagonal
  paramDict['diagonal_direction'] = border.diagonal_direction
  paramDict['vertical'] = border.vertical
  paramDict['horizontal'] = border.horizontal
  paramDict['diagonalUp'] = border.diagonalUp
  paramDict['diagonalDown'] = border.diagonalDown
  paramDict['outline'] = border.outline
  paramDict['start'] = border.start
  paramDict['end'] = border.end
  if not paramName in paramDict:
    raise Exception(f"Bad parameter name provided to updateBorderSide: {paramName}")
  paramDict[paramName] = paramValue
  return pxl_sty.Border(**paramDict)
  
def updateBorderAround(sht: pxl_sht.Worksheet,
                    cellRange: pxl_rng.CellRange, 
                    border: pxl_sty.Border):
  def setBorders(coordList: typing.Tuple[typing.Tuple[int, int]],
                 paramValue: pxl_sty.Side,
                 paramName: str):
    for coords in coordList:
      cell: pxl_cell.Cell = sht.cell(row=coords[0], column=coords[1])  
      thisBorder: pxl_sty.Border = updateBorderParameter(
                  border=cell.border,
                  paramValue=paramValue,
                  paramName=paramName)
      cell.border = thisBorder
  coordList = cellRange.top
  setBorders(coordList = coordList,
             paramValue = border.top,
             paramName = 'top')
  coordList = cellRange.bottom
  setBorders(coordList = coordList,
             paramValue = border.bottom,
             paramName = 'bottom')
  coordList = cellRange.left
  setBorders(coordList = coordList,
             paramValue = border.left,
             paramName = 'left')
  coordList = cellRange.right
  setBorders(coordList = coordList,
             paramValue = border.right,
             paramName = 'right')

def updatePatternFillParameter(patternFill: pxl_sty.PatternFill,
                               paramValue: any,
                               paramName: str,
                               ifNone: bool = False) -> pxl_sty.Border:
  paramDict: typing.Dict = {}
  paramDict['patternType'] = patternFill.patternType
  paramDict['fgColor'] = patternFill.fgColor
  paramDict['bgColor'] = patternFill.bgColor
  paramDict['fill_type'] = patternFill.fill_type
  paramDict['start_color'] = patternFill.start_color
  paramDict['end_color'] = patternFill.end_color
  if not paramName in paramDict:
    raise Exception(f"Bad parameter name provided to updatePatternFillParameter: {paramName}")
  paramDict[paramName] = paramValue
  return pxl_sty.PatternFill(**paramDict)


def updateFill(sht: pxl_sht.Worksheet,
               cellRange: pxl_rng.CellRange, 
               patternFill: pxl_sty.PatternFill,
               ifHasPatternType: bool = True):
  coordsList = cellRange.cells
  for coords in coordsList:
    cell: pxl_cell.Cell = sht.cell(row=coords[0], column=coords[1])
    if cell.fill.patternType is None or ifHasPatternType:
      cell.fill = patternFill

