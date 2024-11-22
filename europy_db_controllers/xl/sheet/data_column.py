from __future__ import annotations

import sys, typing, uuid, datetime, enum

import openpyxl as pxl
from openpyxl import styles as pxl_sty 
from openpyxl.workbook import defined_name as pxl_rng_nm
from openpyxl.worksheet import cell_range as pxl_rng
from openpyxl.cell import cell as pxl_cell
from openpyxl.worksheet import worksheet as pxl_sht


from europy_db_controllers import _controller_base
from europy_db_controllers.xl.sheet import data_block
from europy_db_controllers.xl.sheet import utils as sht_ut
from europy_db_controllers.xl.validation import validation_column as io_val_col
from europy_db_controllers.xl.validation import specific_validations as val_spec
from europy_db_controllers.xl.validation import specific_number_format as frm_spec

if typing.TYPE_CHECKING:
  from europy_db_controllers.xl import row_control
  from europy_db_controllers.xl import col_control


class DataColumn():
  _numberOfRows = 500
  _numberOfRowsPricing = 50000
  _backgroundColor = pxl_sty.Color(rgb="99EBE234")
  _patternFill = pxl_sty.PatternFill(start_color="F6F2A0", end_color="F6F2A0", fill_type = 'solid')
  _borderStyle1 = pxl_sty.Side(style='thin', color='000000')
  _borderStyle2 = pxl_sty.Side(style='dotted', color='000000')
  _border = pxl_sty.Border(left=_borderStyle1, right=_borderStyle1,
                           top=None, bottom=_borderStyle2,
                           vertical=None, horizontal=None)
  _labelAlignment = pxl_sty.Alignment(textRotation=90, horizontal = 'center')
  def __init__(self,
               label: str,
               subControllerKey: _controller_base.BaseControllerKeyEnum,
               controllerKeyEnum: enum.Enum,
               validation: io_val_col.ValidationColumn,
               columnNumber: int,
               rowControl: 'row_control.RowControl',
               colControl: 'col_control.ColControl',
               unique: bool, 
               sqlalchemyDataType: str
               ) -> None:
    self.label: str = label
    self.subControllerKey = subControllerKey
    self.controllerKeyEnum = controllerKeyEnum
    self.validation = validation
    self._colNo: int = columnNumber
    self.unique = unique
    self.sqlalchemyDataType = sqlalchemyDataType
    self._rowControl: 'row_control.RowControl' = rowControl
    self._colControl: 'col_control.ColControl' = colControl

    self.sht: pxl_sht.Worksheet = None

  @property
  def columnNumber(self) -> int:
    return self._colNo
  @property
  def rowControl(self) -> 'row_control.RowControl':
    return self._rowControl
  @property  
  def colControl(self) -> 'col_control.ColControl':
    return self._colControl
  @property
  def columnLabelRow(self) -> int:
    return self._rowControl.columnLableRow
  @property
  def labelCell(self) -> pxl_cell.Cell:
    return self.sht.cell(self.columnLabelRow, self.columnNumber)


  @property
  def dataRangeDelimiters(self) -> typing.Dict[str, int]:
    result = dict[str, int]()
    result['min_row'] = self.columnLabelRow + 1
    if self.subControllerKey == self.controllerKeyEnum.ASSET_PRICING:
      result['max_row'] = self._rowControl.lastDataRow + self._numberOfRowsPricing
    else:
      result['max_row'] = self._rowControl.lastDataRow + self._numberOfRows
    result['min_col'] = self.columnNumber
    result['max_col'] = self.columnNumber
    return result
  @property
  def dataRange(self) -> pxl_rng.CellRange:
    return pxl_rng.CellRange(**self.dataRangeDelimiters)
  @property
  def quotedSheetTitle(self):
    return pxl.utils.cell.quote_sheetname(self.sht.title)
  @property 
  def localDataRangeName(self) -> str:
    return f"{self._colControl.tableName}_{self.label}"
  @property
  def globalDataRangeName(self) -> str:
    return f"{self.quotedSheetTitle}!{self.localDataRangeName}"
  @property 
  def fullDataRangeAddress(self) -> str:
    coord = pxl.utils.absolute_coordinate(self.dataRange.coord)
    return f"{self.quotedSheetTitle}!{coord}"


  def setupLabel(self,
                 sht: pxl_sht.Worksheet) -> None:
    self.sht = sht
    self.labelCell.border = self._colControl._border
    self.labelCell.fill = self._colControl._patternFill
    self.labelCell.alignment = self._labelAlignment
    self.labelCell.value = self.label

  def writeValue(self,
                 cellValue: any,
                 dataBlock: data_block.DataBlock) -> None:
    cellAddress = sht_ut.getCellAddress(row = dataBlock.dataRow,
                                        col = self.columnNumber)
    cell = self.sht[cellAddress] 
    if isinstance(cellValue, uuid.UUID):
      cellValue = str(cellValue)
    elif isinstance(cellValue, datetime.datetime):
      cellValue = cellValue.strftime('%Y-%m-%d %H:%M:%S')
    cell.value = cellValue
    cell.fill = self._patternFill
    cell.border = self._border

  def setFormatsAndValidations(self,
                               sht: pxl_sht.Worksheet):
    self.sht = sht
    definedName = pxl_rng_nm.DefinedName(name = self.localDataRangeName,
                                         attr_text = self.fullDataRangeAddress)
    self.sht.defined_names.add(definedName)
    # identify numberFormat:
    numberFormatParameter = frm_spec.getNumberFormatParameters(sqlalchemyDataType=self.sqlalchemyDataType)
    for row in self.sht.iter_rows(**self.dataRangeDelimiters):
      for rowCell in row:
        # rowCell: pxl_sht.Cell = rowCellTuple[0]
        rowCell.fill = self._patternFill
        rowCell.border = self._border
        if not numberFormatParameter is None:
          rowCell.number_format = numberFormatParameter
    if self.validation is None:
      dataValidation = val_spec.getValidation(
                          sht = self.sht,
                          cellRow = self._rowControl.firstDataRow,
                          cellCol = self.columnNumber,
                          cellRange=self.dataRange,
                          unique=self.unique,
                          sqlalchemyDataType = self.sqlalchemyDataType)
      if not dataValidation is None:
        dataValidation.add(self.dataRange.coord)
        self.sht.add_data_validation(dataValidation)
    if not self.validation is None:
      validationSourceSheetTitle = None
      validationSourceCellRange = None
      if self.validation.isInternalValidation:
        validationSourceSheetTitle = self.validation.validationLocator[1]
        validationSourceSheet = self.validation.wkb[validationSourceSheetTitle]
        validationSourceRangeName = f"{validationSourceSheetTitle}_name"
        validationSourceRange = validationSourceSheet.defined_names[validationSourceRangeName]
        for destination in validationSourceRange.destinations:
          validationSourceCellRange = pxl_rng.CellRange(range_string=destination[1])
      dataValidation = self.validation.dataValidation(quotedSheetTitle = validationSourceSheetTitle,
                                                      cellRange = validationSourceCellRange)
      dataValidation.add(self.dataRange.coord)
      self.sht.add_data_validation(dataValidation)

  def identify(self,
               sht: pxl_sht.Worksheet) -> None:
    self.sht = sht
    if not self.localDataRangeName in self.sht.defined_names:
      raise Exception("Missing defined name on ioWorksheet.\n" + \
                      f"Defined name missing: {self.localDataRangeName}\n" + \
                      f"On ioWorksheet: {self.sht.title}")
    definedName = self.sht.defined_names[self.localDataRangeName]
    dest = definedName.destinations
    for title, coord in dest:
      lastCoord = coord.split(':')[1]
      lastRow = self.sht[lastCoord].row
      self.rowControl.updateDataRow(lastRow)

  def getValueOfRow(self, 
                    row: int) -> any:
    coords = sht_ut.getCellAddress(row = row,
                                   col = self.columnNumber)
    return self.sht[coords].value
  def getLabelAndValueOfRow(self,
                            row: int) -> typing.Tuple[str, any]:
    return (self.label, self.getValueOfRow(row = row))

    

