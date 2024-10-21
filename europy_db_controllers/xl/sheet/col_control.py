from __future__ import annotations

import sys, typing

import sqlalchemy
from sqlalchemy import orm as sqlalchemy_orm
from sqlalchemy import schema as sqlalchemy_schema
from sqlalchemy.ext import declarative as sqlalchemy_decl

import openpyxl as pxl
from openpyxl import styles as pxl_sty 
from openpyxl.worksheet import worksheet as pxl_sht
from openpyxl.cell import cell as pxl_cell
from openpyxl.worksheet import cell_range as pxl_rng

sys.path.insert(0, '..\..\..')

from europy_db_controllers import _controller_base
from europy_db_controllers.entity_capsules import _capsule_utils, _capsule_base
from europy_db_controllers.xl.sheet import row_control, data_column, utils, \
                                     delete_cntr_column, col_control_data
from europy_db_controllers.xl.sheet import data_block
from europy_db_controllers.xl.validation import validation_sht as io_val
from europy_db_controllers.xl.validation import validation_column as io_val_col

if typing.TYPE_CHECKING:
  from europy_db_controllers import controller

debug_rec_count: int = 0

DC = typing.TypeVar("DC", bound=data_column.DataColumn)
CT = typing.TypeVar("CT", bound=_capsule_base.CapsuleBase)


class ColControl():
  _backgroundColor = pxl_sty.Color(rgb="EBD334")
  _patternFill = pxl_sty.PatternFill(start_color="EBD334", end_color="EBD334", fill_type = 'solid')
  _borderStyle = pxl_sty.Side(style='thick', color='000000')
  _border = pxl_sty.Border(left=_borderStyle, right=_borderStyle,
                           top=_borderStyle, bottom=_borderStyle,
                           vertical=None, horizontal=None)
  _labelAlignment = pxl_sty.Alignment(horizontal = 'center') 
  _labelFont = pxl_sty.Font(size = 14, bold = True)

  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  # Properties et al
  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

  @property
  def isList(self) -> bool:
    return self._isList
  @property
  def tableName(self) -> str:
    return str(self._capsuleKey)
  @property
  def label(self) -> str:
    if self._relationshipKey is None: return self.tableName
    return self._relationshipKey
  @property
  def firstRow(self) -> int:
    return self.__getHeadDepth()
  @property
  def lastRow(self) -> int:
    return self._rowControl.columnLableRow - 1
  @property
  def lastCol(self) -> int:
    return self.firstCol + self._width() - 1

  def _width(self) -> int:
    result = len(self.columns) 
    for subColControl in self.subColControls.values():
      result = result + subColControl._width()
    return result

  @property
  def labelCellAddress(self) -> str:
    return utils.getCellAddress(
                    row = self.firstRow, 
                    col = self.firstCol)
  @property
  def labelRangeAddress(self) -> str:
    return utils.getRangeAddress(
                    luRow = self.firstRow, luCol = self.firstCol,
                    rlRow = self.firstRow, rlCol = self.lastCol)
  @property
  def labelCell(self) -> pxl_cell.Cell:
    return self.sht.cell(row = self.firstRow, column = self.firstCol)
  @property
  def fullRange(self) -> pxl_rng.CellRange:
    return pxl_rng.CellRange(title=self.sht.title, 
                             min_row=self.firstRow, min_col=self.firstCol,
                             max_row=self.lastRow, max_col=self.lastCol)

  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  # Shared internal functions 
  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  def __isMain(self) -> bool:
    return self._parentColControl is None
  def __getHeadDepth(self) -> int:
    if self.__isMain(): return 1
    else: return self._parentColControl.__getHeadDepth() + 1
  def __updateRowControl(self):
    headDepth = self.__getHeadDepth()
    self._rowControl.updateColumnLabelRow(
                  columnLableRow=headDepth + 1)
  def __getRelationshipFeatures(self,
                                relationship: sqlalchemy_orm.Relationship,
                                ) -> typing.Tuple[bool, 
                                                  bool,
                                                  bool, 
                                                  bool]:
    def relationshipIsPartOfListOf() -> bool:
      if hasattr(self._sqlalchemyType, '_is_part_of_list_of'):
        if relationship.key in self._sqlalchemyType._is_part_of_list_of:
          return True
      return False
    isDisplayList = _capsule_utils.isDisplayList(
                sqlalchemyTableType = self._sqlalchemyType,
                relationship = relationship)
    isExcludedFromJson = relationship.key in self._sqlalchemyType._exclude_from_json
    relationshipDecl = relationship.mapper.class_
    relationshipDeclHasName = hasattr(relationshipDecl, 'name')
    isPartOfListOf = relationshipIsPartOfListOf()
    return(isDisplayList, isExcludedFromJson, isPartOfListOf, relationshipDeclHasName)
  def __getRelationshipDefinitions(self,
                                   relationship: sqlalchemy_orm.Relationship
                                   ) -> typing.Tuple[str, 
                                                     sqlalchemy_decl.DeclarativeMeta,
                                                     sqlalchemy_schema.Table, 
                                                     type[CT]]:
    relName: str = relationship.key
    relDecl: sqlalchemy_decl.DeclarativeMeta = relationship.mapper.class_
    relDeclTable = relDecl.__table__
    relCapsuleTypeName = _capsule_utils.getCapsuleClassName(table = relDeclTable)
    relCapsuleType = getattr(capsules, relCapsuleTypeName)
    return (relName, relDecl, relDeclTable, relCapsuleType)
  def __getRelationshipDefinitionsOfColumn(self,
                                           column: sqlalchemy_schema.Column
                                           ) -> typing.Tuple[str, 
                                                             sqlalchemy_orm.Relationship, 
                                                             sqlalchemy_decl.DeclarativeMeta,
                                                             sqlalchemy_schema.Table, 
                                                             type[CT]]:
    relName: str = _capsule_utils.getRelationshipNameOfColumn(column=column)
    rel: sqlalchemy_orm.Relationship = self._relationships[relName]
    relName, relDecl, relDeclTable, relCapsuleType = \
            self.__getRelationshipDefinitions(relationship=rel)
    return (relName, rel, relDecl, relDeclTable, relCapsuleType)
  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  # INIT: Define structure of sheet content   
  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  def __addDeleteControlColumn(self) -> None:
    dCL = delete_cntr_column.DeleteControlColumn(columnNumber = self._width() + self.firstCol,
                                                 subControllerKey = self.subControllerKey,
                                                 rowControl = self._rowControl,
                                                 colControl = self)
    self.columns[dCL.label] = dCL
  def __addSingleRelationship(self,
                              column: sqlalchemy_schema.Column) -> None:
    relationshipName, relationship, relationshipDecl, \
    relationshipDeclTable, relationshipCapsuleType = \
                self.__getRelationshipDefinitionsOfColumn(column = column)
    isExcludedFromJson, isPartOfListOf, relationshipDeclHasName = \
                self.__getRelationshipFeatures(relationship = relationship)[1:]
    # If the relationship is excluded from json but has a name, then the 
    #   relationship's 'name' field of the capsule is added as a Column 
    #   to the ColControl 
    #   The 'isPartOfListOf' avoids the inclusion of such 'name'-column in case of
    #   the current capsule is part of a list of the parent ColControl (is redundant
    #   information in such case).
    if isExcludedFromJson:
      if relationshipDeclHasName and not isPartOfListOf:
        relationshipNameAttrName = _capsule_utils.getRelationshipNameFieldOfColumn(column=column)
        validation = self.validations.getValidationOfCapsuleKey(capsuleKey = relationshipCapsuleType._key())
        self._addColumn(label = relationshipNameAttrName,
                        validation = validation,
                        unique=False,
                        sqlalchemyDataType = "") # omit type validation as validated by source
    else:
      self._addSubColControl(capsuleType = relationshipCapsuleType,
                             isList = False,
                             relationshipKey = relationshipName)
  def __addListRelationship(self,
                            relationship: sqlalchemy_orm.Relationship) -> None:
    isDisplayList, isExcludedFromJson = self.__getRelationshipFeatures(relationship = relationship)[:2]
    if not (isDisplayList or isExcludedFromJson):
      relationshipName, relationshipDecl, \
      relationshipDeclTable, relationshipCapsuleType = \
                self.__getRelationshipDefinitions(relationship = relationship)  
      self._addSubColControl(capsuleType = relationshipCapsuleType,
                             isList = True,
                             relationshipKey = relationshipName)
  def __init__(self,
               subControllerKey: _controller_base.ControllerKeyEnum,
               capsuleType: type[CT],
               rowControl: row_control._rowControl,
               validations: io_val.ValidationSheet,
               firstCol: int = 1,
               parentColControl: ColControl = None,
               isList: bool = False,
               relationshipKey: str = None) -> None:
    
    self.subControllerKey = subControllerKey
    self._capsuleType: type[CT] = capsuleType
    self._sqlalchemyType: sqlalchemy_decl.DeclarativeMeta = self._capsuleType.sqlalchemyTableType
    self._table: sqlalchemy_schema.Table = self._sqlalchemyType.__table__
    self._capsuleKey = self._capsuleType._key()    
    self._relationships = self._sqlalchemyType.__mapper__.relationships

    self.firstCol: int = firstCol
    self.validations = validations
    self.subColControls: typing.OrderedDict[str, ColControl] = {}
    self.columns: typing.OrderedDict[str, DC] = {}

    self._isList: bool = isList
    self._rowControl: row_control.RowControl = rowControl
    self._parentColControl: ColControl = parentColControl
    self._relationshipKey: str = relationshipKey

    self.sht: pxl_sht.Worksheet = None

    self.__addDeleteControlColumn()
    for column in self._table.columns:
      columnName = column.name
      if columnName in self._sqlalchemyType._changeTrackFields:
        pass # internal change control only
      elif _capsule_utils.isRelationshipIdColumn(column=column):
        self.__addSingleRelationship(column = column)
      else:
        self._addColumn(label = columnName,
                        validation=None,
                        unique = True if columnName == 'id' else column.unique,
                        sqlalchemyDataType = str(column.type))
    for relationship in self._relationships:
      if relationship.uselist:
        self.__addListRelationship(relationship = relationship)
    self.__updateRowControl()
  def _addColumn(self, 
                 label: str,
                 validation: io_val_col.ValidationColumn,
                 unique: bool,
                 sqlalchemyDataType: str):
    column = data_column.DataColumn(label = label,
                                    subControllerKey = self.subControllerKey,
                                    validation=validation,
                                    columnNumber = self._width() + self.firstCol,
                                    rowControl = self._rowControl,
                                    colControl = self,
                                    unique = unique,
                                    sqlalchemyDataType = sqlalchemyDataType)
    self.columns[column.label] = column
  def _addSubColControl(self,
                        capsuleType: type[CT],
                        isList: bool = False,
                        relationshipKey: str = None
                        ) -> None:  
    subColControl = ColControl(subControllerKey = self.subControllerKey,
                               capsuleType = capsuleType,
                               rowControl = self._rowControl,
                               validations = self.validations,
                               firstCol = self.lastCol + 1,
                               parentColControl = self,
                               isList = isList,
                               relationshipKey=relationshipKey)
    self.subColControls[subColControl.label] = subColControl

  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  # DOWNLOAD: Setup sheet and write data  
  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  def _setupLabelRange(self) -> None:
    labelCell: pxl_sht.Cell = self.labelCell
    labelCell.alignment = self._labelAlignment
    labelCell.font = self._labelFont
    labelCell.fill = self._patternFill
    self.sht[self.labelCellAddress] = self.label
    self.sht.merge_cells(self.labelRangeAddress)
  def _setupFullRange(self) -> None:
    fullRange = self.fullRange
    utils.updateFill(sht = self.sht,
                     cellRange=fullRange,
                     patternFill=self._patternFill,
                     ifHasPatternType=False)    
    utils.updateBorderAround(sht = self.sht,
                             cellRange = fullRange,
                             border=self._border) 
  def setupLabels(self,
                  sht: pxl_sht.Worksheet) -> None:
    self.sht = sht
    self._setupFullRange()
    self._setupLabelRange()
    for dataColumn in self.columns.values():
      dataColumn.setupLabel(sht = self.sht)
    for subColControl in self.subColControls.values():
      subColControl.setupLabels(sht = self.sht)
  def writeValues(self,
                  colControlDict: typing.Dict,
                  dataBlock: data_block.DataBlock):
    for dataColumn in self.columns.values():
      if isinstance(dataColumn, delete_cntr_column.DeleteControlColumn): continue   
      cellValue = colControlDict[dataColumn.label]
      dataColumn.writeValue(cellValue = cellValue,
                            dataBlock=dataBlock)
    for subColControl in self.subColControls.values():
      subDataBlock = dataBlock.nextSubBlock(isList=subColControl.isList,
                                         minCol=subColControl.firstCol,
                                         maxCol=subColControl.lastCol,
                                         colBlockName=subColControl.label,
                                         colBlockTableName=subColControl.tableName)
      if not subColControl.label in colControlDict: continue
      thisColControlDict = colControlDict[subColControl.label]
      if thisColControlDict is None: continue
      if len(thisColControlDict) == 0: continue
      if subColControl.isList:
        listCount: int = 0
        while listCount in thisColControlDict:
          thisDataBlock = subDataBlock.nextListElement()
          thisColControlItemDict = thisColControlDict[listCount]
          subColControl.writeValues(colControlDict = thisColControlItemDict,
                                    dataBlock=thisDataBlock)
          listCount += 1
      else:
        subColControl.writeValues(colControlDict = thisColControlDict,
                                  dataBlock=subDataBlock)
    self._rowControl.updateDataRow(dataRow = dataBlock.maxRow)
  def setFormatsAndValidations(self,
                               sht: pxl_sht.Worksheet):
    # FIXME: set cell formats
    self.sht = sht
    for dataColumn in self.columns.values():
      dataColumn.setFormatsAndValidations(sht = self.sht)
    for subColControl in self.subColControls.values():
      subColControl.setFormatsAndValidations(sht = self.sht)

  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  # UPLOAD: Data content identification    
  # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
  def getColControlRowDelimiters(self,
                                 row: int) -> typing.Dict[str, int]:
    result = dict[str, int]()
    result['min_row'] = row
    result['max_row'] = row
    result['min_col'] = self.firstCol
    result['max_col'] = len(self.columns) + self.firstCol - 1
    return result
  def isEmptyColControlRow(self,
                           row: int) -> bool:
    for cellTuple in self.sht.iter_cols(**self.getColControlRowDelimiters(row = row)):
      if cellTuple[0].value is not None: return False
    return True
  def getColControlContent(self,
                           row: int) -> typing.Dict[str, any]:
    result = dict[str, any]()
    for column in self.columns.values():
      key, value = column.getLabelAndValueOfRow(row = row)
      result[key] = value
    return result
  def identify(self,
               sht: pxl_sht.Worksheet) -> None:
    self.sht = sht
    for dataColumn in self.columns.values():
      dataColumn.identify(sht = self.sht)
    for subColControl in self.subColControls.values():
      subColControl.identify(sht = self.sht)

  def _identifyData(self,
                    dataBlock: data_block.DataBlock) -> None:
    if dataBlock.isList:
      for row in range(dataBlock.dataRow, dataBlock.maxRow + 1):
        # if doPrint: print(f"  row: {row} - isEmpty: {self.isEmptyColControlRow(row = row)}")  
        # if not self.isEmptyColControlRow(row = row) and doPrint:
        #   for cellTuple in self.sht.iter_cols(**self.getColControlRowDelimiters(row = row)):
        #     print(f"    row: {cellTuple[0].row} - column: {cellTuple[0].column} - cellTuple[0].value: {cellTuple[0].value}")
        if self.isEmptyColControlRow(row = row): continue
        valueDict = self.getColControlContent(row = row)
        colControlData = col_control_data.ColControlData(colControlLabel = self.label,
                                                        sht = self.sht,
                                                        row = row,
                                                        valueDict=valueDict)
        colControlData.ensureConsistency()
        nextListBlock = dataBlock.nextListElement(dataRow=row)
        nextListBlock.colControlData = colControlData
      for listElement in dataBlock.listElements:
        for subColControl in self.subColControls.values():
          subDataBlock = listElement.nextSubBlock(isList=subColControl.isList,
                                                 minCol=subColControl.firstCol,
                                                 maxCol=subColControl.lastCol,
                                                 colBlockName=subColControl.label,
                                                 colBlockTableName=subColControl.tableName)
          subColControl._identifyData(dataBlock=subDataBlock)
    else:
      for row in range(dataBlock.parent.dataRow, dataBlock.parent.maxRow + 1):
        valueDict = self.getColControlContent(row = dataBlock.dataRow)
        colControlData = col_control_data.ColControlData(colControlLabel = self.label,
                                                          sht = self.sht,
                                                          row = row,
                                                          valueDict=valueDict)
        colControlData.ensureConsistency()
        if row == dataBlock.dataRow:
          dataBlock.colControlData = colControlData
        else:
          # if not a list, all rows must be empty
          colControlData.ensureEmpty
        for subColControl in self.subColControls.values():
          subDataBlock = dataBlock.nextSubBlock(isList=subColControl.isList,
                                                minCol=subColControl.firstCol,
                                                maxCol=subColControl.lastCol,
                                                colBlockName=subColControl.label,
                                                colBlockTableName=colControlData.tableName,
                                                dataRow = listElement.dataRow)
          subColControl._identifyData(dataBlock=subDataBlock)

    # doPrint = False
    # if self.sht.title == 'client': doPrint = True
    # if doPrint and dataBlock.parent is None:
    #   print(f"[0] dataBlock - dataBlock.colBlockName: {dataBlock.colBlockName} - dataBlock.isList: {listElement.isList}")
    #   print(f"    dataBlock.dataRow: {dataBlock.dataRow} - dataBlock.maxRow: {dataBlock.maxRow}") 
    #   print(f"    values (is None): {dataBlock.colControlData is None}")
    #   for listElement in dataBlock.listElements:
    #     print(f"  [1] listElement - listElement.colBlockName: {listElement.colBlockName} - listElement.isList: {listElement.isList}")
    #     print(f"      listElement.dataRow: {listElement.dataRow} - listElement.maxRow: {listElement.maxRow}") 
    #     print(f"      len(listElement.subBlocks): {len(listElement.subBlocks)}")
    #     print(f"      values: {str(listElement.colControlData.valueDict)}")
    #     subBlock = listElement.subBlocks[0]
    #     print(f"    [2] subBlock - subBlock.colBlockName: {subBlock.colBlockName} - subBlock.isList: {subBlock.isList}")
    #     print(f"        subBlock.dataRow: {subBlock.dataRow} - subBlock.maxRow: {subBlock.maxRow}") 
    #     print(f"        values (is None): {subBlock.colControlData is None}")
    #     for secondLevelElement in subBlock.listElements:
    #       print(f"      [3] secondLevelElement - secondLevelElement.colBlockName: {secondLevelElement.colBlockName} - secondLevelElement.isList: {listElement.isList}")
    #       print(f"          secondLevelElement.dataRow: {secondLevelElement.dataRow} - secondLevelElement.maxRow: {secondLevelElement.maxRow}") 
    #       print(f"          values: {str(secondLevelElement.colControlData.valueDict)}")
    #   print("")
  def parentPartOfListOfChildDefs(self,
                                  relationshipSqlaTable: sqlalchemy_decl.DeclarativeMeta,
                                  relationshipTable: sqlalchemy.Table) -> typing.Tuple[bool, str]:
    for column in relationshipTable.columns:
      if _capsule_utils.isRelationshipIdColumn(column=column):
        parentRelationshipName: str = _capsule_utils.getRelationshipNameOfColumn(column=column)
        parentRelationship = relationshipSqlaTable.__mapper__.relationships[parentRelationshipName]
        parentRelationshipSqlaTable = parentRelationship.mapper.class_
        parentRelationshipTable = parentRelationshipSqlaTable.__table__
        isPartOfListOf = False
        if hasattr(relationshipSqlaTable, '_is_part_of_list_of'):
          if parentRelationship.key in relationshipSqlaTable._is_part_of_list_of:
            isPartOfListOf = True and hasattr(parentRelationshipSqlaTable, 'name')
        parentNameAttributeField = _capsule_utils.getRelationshipNameFieldOfColumn(column = column)
        if isPartOfListOf and parentRelationshipTable.name == self._table.name: return (True, parentNameAttributeField)
    return (False, None)
        

  def toDict(self,
             dataEntry: data_block.DataBlock) -> typing.Dict[str, any]:
    result: typing.Dict[str, any] = dict[str, any]()
    if not dataEntry.colControlData.hasDeleteMarker:
      for column in self._table.columns:
        columnName = column.name
        if columnName in self._sqlalchemyType._changeTrackFields:
          pass # internal change control only
        elif _capsule_utils.isRelationshipIdColumn(column=column):
          relationshipName, relationship = self.__getRelationshipDefinitionsOfColumn(column = column)[:2]
          isExcludedFromJson, relationshipDeclHasName = \
                      self.__getRelationshipFeatures(relationship = relationship)[1:4:2]
          if isExcludedFromJson:
            if relationshipDeclHasName:
              # no control for 'isPartOfListOf' here. If 'isPartOfListOf', the value must be added
              #     to the dataEntry.colControlData while processing the parent, see below [#add name of parent].
              relationshipNameAttrName = _capsule_utils.getRelationshipNameFieldOfColumn(column=column)
              value = dataEntry.colControlData.getValue(relationshipNameAttrName)
              result[relationshipNameAttrName] = value              
          else:
            relationshipDataEntry = dataEntry.getSubBlockOfName(colBlockName = relationshipName)
            subColControl = self.subColControls[relationshipName]
            result[relationshipName] = subColControl.toDict(dataEntry = relationshipDataEntry)
        else:
          value = dataEntry.colControlData.getValue(columnName)
          result[columnName] = value
      for relationship in self._relationships:
        if relationship.uselist:
          isDisplayList, isExcludedFromJson = self.__getRelationshipFeatures(relationship = relationship)[:2]
          if not (isDisplayList or isExcludedFromJson):

            # relationshipName, relationshipTable = \
            #               self.__getRelationshipDefinitions(relationship = relationship)[0:3:2]  
            relationshipName, relationshipSqlaTable, relationshipTable = \
                          self.__getRelationshipDefinitions(relationship = relationship)[0:3:1]  

            relationshipDataList = dataEntry.getSubBlockOfName(colBlockName = relationshipName)
            subColControl = self.subColControls[relationshipName]
            # [#Evaluate if parent name to add to dict] identify whether the relationship entity has a link to this parent
            parentPartOfListOfChildDefs = self.parentPartOfListOfChildDefs(relationshipSqlaTable = relationshipSqlaTable,
                                                                           relationshipTable = relationshipTable)
            # relationshipDataEntry is a list DataBlock
            entryCount = 0
            relationshipDataDict: typing.Dict[int, any] = dict[int, any]()
            for relationshipDataEntry in relationshipDataList.listElements:
              # [#add name of parent] add the parent name attribute field of the current entity to the dictionary of child values
              #    depending on the evaluation of the relationship, see above [#Evaluate if parent name to add to dict]
              if not relationshipDataEntry.colControlData.isEmpty:
                if parentPartOfListOfChildDefs[0]:
                  parentName = dataEntry.colControlData.getValue('name')
                  relationshipDataEntry.colControlData.valueDict[parentPartOfListOfChildDefs[1]] = parentName
                relationshipDict = subColControl.toDict(dataEntry = relationshipDataEntry)
                relationshipDataDict[entryCount] = relationshipDict
                entryCount += 1
            result[relationshipName] = relationshipDataDict
    return result

  def getDeleteDict(self,
                    dataEntry: data_block.DataBlock) -> typing.Dict[str, list]:   
    result: typing.Dict[str, list] = dict[str, list]()
    def __ensureListInResult(colBlockName: str) -> None:
      if not colBlockName in result:
        result[colBlockName] = []
    def deleteIfMarked(colControlData: col_control_data.ColControlData,
                       colBlockName: str):
      if colControlData.hasDeleteMarker:
        if colControlData.hasId:
          __ensureListInResult(colBlockName=colBlockName)
          result[colBlockName].append(colControlData.getId())
        return True
      return False
         
      
    if dataEntry.isList:
      colBlockName = dataEntry.colBlockName
      for listElement in dataEntry.listElements:
        colControlData = listElement.colControlData
        deleted = deleteIfMarked(colControlData = colControlData,
                                 colBlockName = colBlockName)
        # no further delete required - cascade delete must be defined on relationships
        if not deleted:
          subResult = listElement.getDeleteDict()
          # iterate the subBlocks
          pass
    else:
      colControlData = dataEntry.colControlData
      deleted = deleteIfMarked(colControlData = colControlData,
                               colBlockName = colBlockName)
      # no further delete required - cascade delete must be defined on relationships
    return result

        
